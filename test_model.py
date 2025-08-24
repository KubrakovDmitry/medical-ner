"""Программа для тестирования языковых моделей"""

from transformers import AutoModelForTokenClassification, AutoTokenizer
import torch
from openpyxl import Workbook

from auxiliary_module import (load_json_data, merge_tokens_for_bert,
                              merge_tokens_for_t5, convert_to_dataframe)


wb = Workbook()
# делаем единственный лист активным 
ws = wb.active

texts = [
  ('Периндоприл оказывает сосудорасширяющее действие , '
   'способствует восстановлению эластичности крупных артерий и структуры сосудистой '
   'стенки мелких артерий , а также уменьшает гипертрофию левого желудочка . '
   'Одновременное применение тиазидных диуретиков group усиливает выраженность '
   'антигипертензивного эффекта. Сердечная недостаточность Периндоприл prepare '
   'нормализует работу сердца , снижая преднагрузку и постнагрузку. '
   'У пациентов с хронической сердечной недостаточностью (ХСН) , '
   'получавших периндоприл , было выявлено снижение давления наполнения в левом '
   'и правом желудочках сердца, снижение ОПСС, повышение сердечного выброса '
   'и увеличение сердечного индекса .'),
  ]


BIO_mark_sent_v2_rubert_base_cased_model_names = [
  'models\\model_data_bio_4_rubert-base-cased_6', 
]

BIO_mark_sent_v2_RuBioBERT_model_names = [
  'models\\model_data_bio_4_RuBioBERT_6',]


BIO_mark_sent_v2_rut5_base_model_names = [
  'models\\model_data_bio_4_rut5-base_6',
]


BIO_mark_sent_v2_rut5_base_multitask_model_names = [
  'models\\model_data_bio_4_rut5-base-multitask_6',
]

annotation = 'content\\data_bio_4.json'

data = load_json_data(annotation)
df = convert_to_dataframe(data)

label_list = list()
for i in range(df.shape[0]):
    label_list.extend((list(df.tags.iloc[i])))

label_list = set(label_list)


first_row = 1
count = 0
for text in texts:

  count += 1

  column_token_rubert = 1
  column_ner_rubert = 2
  column_token_rubiobert = 3
  column_ner_rubiobert = 4

  column_token_rut5 = 5
  column_ner_rut5 = 6
  column_token_rut5_multitask = 7
  column_ner_rut5_multitask = 8
  
  
  for model_name1, model_name2, model_name3, model_name4 in zip(BIO_mark_sent_v2_rubert_base_cased_model_names, 
                                                                BIO_mark_sent_v2_RuBioBERT_model_names,
                                                                BIO_mark_sent_v2_rut5_base_model_names,
                                                                BIO_mark_sent_v2_rut5_base_multitask_model_names):
    number_row = 2
    # Загрузка токенизатора и модели
    tokenizer1 = AutoTokenizer.from_pretrained(model_name1, trust_remote_code=True)
    model1 = AutoModelForTokenClassification.from_pretrained(model_name1, num_labels=len(label_list), trust_remote_code=True)

    inputs1 = tokenizer1(text, return_tensors="pt")

    with torch.no_grad():
      logits = model1(**inputs1).logits

    predictions1 = torch.argmax(logits, dim=2)
    predicted_token_class1 = [model1.config.id2label[t.item()] for t in predictions1[0]]

    tokenizer2 = AutoTokenizer.from_pretrained(model_name2, trust_remote_code=True)
    model2 = AutoModelForTokenClassification.from_pretrained(model_name2, num_labels=len(label_list), trust_remote_code=True)

    inputs2 = tokenizer2(text, return_tensors="pt")

    with torch.no_grad():
      logits = model2(**inputs2).logits

    predictions2 = torch.argmax(logits, dim=2)
    predicted_token_class2 = [model2.config.id2label[t.item()] for t in predictions2[0]]
    

    # Загрузка токенизатора и модели
    tokenizer3 = AutoTokenizer.from_pretrained(model_name3, trust_remote_code=True)
    model3 = AutoModelForTokenClassification.from_pretrained(model_name3, num_labels=len(label_list), trust_remote_code=True)

    inputs3 = tokenizer3(text, return_tensors="pt")

    with torch.no_grad():
      logits = model3(**inputs3).logits

    predictions3 = torch.argmax(logits, dim=2)
    predicted_token_class3 = [model3.config.id2label[t.item()] for t in predictions3[0]]

    tokenizer4 = AutoTokenizer.from_pretrained(model_name4, trust_remote_code=True)
    model4 = AutoModelForTokenClassification.from_pretrained(model_name4, num_labels=len(label_list), trust_remote_code=True)

    inputs4 = tokenizer4(text, return_tensors="pt")

    with torch.no_grad():
      logits = model4(**inputs4).logits

    predictions4 = torch.argmax(logits, dim=2)
    predicted_token_class4 = [model4.config.id2label[t.item()] for t in predictions4[0]]
    

    ws.cell(row=first_row, column=column_token_rubert, value=model_name1)
    ws.cell(row=first_row, column=column_token_rubiobert, value=model_name2)

    ws.cell(row=first_row, column=column_token_rut5, value=model_name3)
    ws.cell(row=first_row, column=column_token_rut5_multitask, value=model_name4)

    print('column_token_rubert =', column_token_rubert)
    print('column_ner_rubert =', column_ner_rubert)

    print('column_token_rut5 =', column_token_rut5)
    print('column_ner_rut5 =', column_ner_rut5)

    print('column_token_rut5_multitask =', column_token_rut5_multitask)
    print('column_ner_rut5_multitask =', column_ner_rut5_multitask)

    ws.merge_cells(start_row=first_row, start_column=column_token_rubert, end_row=first_row, end_column=column_ner_rubert)
    ws.merge_cells(start_row=first_row, start_column=column_token_rubiobert, end_row=first_row, end_column=column_ner_rubiobert)

    tokens1 = tokenizer1.convert_ids_to_tokens(inputs1["input_ids"][0])
    tokens2 = tokenizer2.convert_ids_to_tokens(inputs2["input_ids"][0])

    ws.merge_cells(start_row=first_row, start_column=column_token_rut5, end_row=first_row, end_column=column_ner_rut5)
    ws.merge_cells(start_row=first_row, start_column=column_token_rut5_multitask, end_row=first_row, end_column=column_ner_rut5_multitask)

    tokens3 = tokenizer3.convert_ids_to_tokens(inputs3["input_ids"][0])
    tokens4 = tokenizer4.convert_ids_to_tokens(inputs4["input_ids"][0])

    tokens1, predicted_token_class1 = merge_tokens_for_bert(tokens1, predicted_token_class1)
    tokens2, predicted_token_class2 = merge_tokens_for_bert(tokens2, predicted_token_class2)
    tokens3, predicted_token_class3 = merge_tokens_for_t5(tokens3, predicted_token_class3)
    tokens4, predicted_token_class4 = merge_tokens_for_t5(tokens4, predicted_token_class4)

    for source_text1, pred1, source_text2, pred2, \
        source_text3, pred3, source_text4, pred4 in zip(tokens1, predicted_token_class1, 
                                                        tokens2, predicted_token_class2,
                                                        tokens3, predicted_token_class3,
                                                        tokens4, predicted_token_class4,
                                                        ):

      ws.cell(row=number_row, column=column_token_rubert, value=source_text1)
      ws.cell(row=number_row, column=column_ner_rubert, value=pred1)
      ws.cell(row=number_row, column=column_token_rubiobert, value=source_text2)
      ws.cell(row=number_row, column=column_ner_rubiobert, value=pred2)

      ws.cell(row=number_row, column=column_token_rut5, value=source_text3)
      ws.cell(row=number_row, column=column_ner_rut5, value=pred3)
      ws.cell(row=number_row, column=column_token_rut5_multitask, value=source_text4)
      ws.cell(row=number_row, column=column_ner_rut5_multitask, value=pred4)

      number_row += 1

    column_token_rubert += 8
    column_ner_rubert += 8
    column_token_rubiobert += 8
    column_ner_rubiobert += 8

    column_token_rut5 += 8
    column_ner_rut5 += 8
    column_token_rut5_multitask += 8
    column_ner_rut5_multitask += 8

  wb.save(f'comperison_BIO_NER_model_{count}.xlsx')

print('Работа программы успешно завершина!')


