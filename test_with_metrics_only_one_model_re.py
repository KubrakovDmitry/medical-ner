from transformers import AutoModelForSequenceClassification, AutoTokenizer
import torch
from openpyxl import Workbook
import pandas as pd
from sklearn.metrics import (accuracy_score,
                             precision_score,
                             recall_score,
                             f1_score,
                             roc_curve,
                             auc,)
import numpy as np
from tqdm import tqdm


TEST_PATH = 'D:\\The job\\NaD\\BIO_NER\\content\\test_relations_3.csv'
NUM_CLASSES = 2
FIRST_ROW = 1
COLUMN_MODEL_NAME = 1
COLUMN_ACCURACY = 2
COLUMN_PRECISION = 3
COLUMN_RECALL = 4
COLUMN_F1 = 5
COLUMN_AUC = 6
MAX_LENGTH = 512
CPU = 'cpu'
test_df = pd.read_csv(TEST_PATH, delimiter='\t')
# Загрузка модели и токенизатора
model_names = [
        'D:\\The job\\NaD\\BIO_NER\\models\\saved_model_content\\changed_data_relations_3_RuBioBERT_3',
        'D:\\The job\\NaD\\BIO_NER\\models\\saved_model_content\\changed_data_relations_3_RuBioBERT_4',
        'D:\\The job\\NaD\\BIO_NER\\models\\saved_model_content\\changed_data_relations_3_RuBioBERT_5',
        'D:\\The job\\NaD\\BIO_NER\\models\\saved_model_content\\changed_data_relations_3_RuBioBERT_6',
        'D:\\The job\\NaD\\BIO_NER\\models\\saved_model_content\\changed_data_relations_3_RuBioBERT_7',
        'D:\\The job\\NaD\\BIO_NER\\models\\saved_model_content\\changed_data_relations_3_RuBioBERT_8',
        'D:\\The job\\NaD\\BIO_NER\\models\\saved_model_content\\changed_data_relations_3_RuBioBERT_9',
        'D:\\The job\\NaD\\BIO_NER\\models\\saved_model_content\\changed_data_relations_3_RuBioBERT_10',
    ]
wb = Workbook()
# делаем единственный лист активным 
ws = wb.active
ws.cell(row=FIRST_ROW, column=COLUMN_MODEL_NAME, value='Название модели')
ws.cell(row=FIRST_ROW, column=COLUMN_ACCURACY, value='Accuracy')
ws.cell(row=FIRST_ROW, column=COLUMN_PRECISION, value='Precision')
ws.cell(row=FIRST_ROW, column=COLUMN_RECALL, value='Recall')
ws.cell(row=FIRST_ROW, column=COLUMN_F1, value='F1')
ws.cell(row=FIRST_ROW, column=COLUMN_AUC, value='AUC')
test_df['type'] = test_df['type'].astype(object)
print('test_df["type"].dtype =', type(test_df['type'][0]))
test_sentences = test_df['left'] + ' ' + test_df['right']
test_sentences = test_sentences.tolist()
test_labels = test_df['type'].tolist()
row_number = 1
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
for model_name in tqdm(model_names, ncols=80):
    row_number += 1
    pred_labels = list()
    probs = []
    tokenizer = AutoTokenizer.from_pretrained(model_name,
                                              trust_remote_code=True)
    model = AutoModelForSequenceClassification.from_pretrained(model_name,
                                                               num_labels=NUM_CLASSES,
                                                               trust_remote_code=True)
    model.to(device)
    # Перевод модели в режим оценки
    model.eval()
    # Токенизация данных
    for test_sentence in test_sentences:
        inputs = tokenizer(test_sentence,
                           padding=True,
                           truncation=True,
                           max_length=MAX_LENGTH,
                           return_tensors="pt")
        inputs = {key: value.to(device) for key, value in inputs.items()}
        # Получение предсказаний
        with torch.no_grad():
            logits = model(**inputs).logits
        predicted_class = torch.argmax(logits, dim=1).item()
        probs.append(torch.softmax(logits, dim=1).cpu().numpy()[0][1])
        pred_labels.append(predicted_class)
    # Вычисление метрик
    accuracy = accuracy_score(test_labels, pred_labels)
    precision = precision_score(test_labels, pred_labels, average='binary')
    recall = recall_score(test_labels, pred_labels, average='binary')
    f1 = f1_score(test_labels, pred_labels, average='binary')

    print(f"Точность (Accuracy): {accuracy}")
    print(f"Точность (Precision): {precision}")
    print(f"Полнота (Recall): {recall}")
    print(f"F1-Score: {f1}")
    ws.cell(row=row_number, column=COLUMN_MODEL_NAME, value=model_name)
    ws.cell(row=row_number, column=COLUMN_ACCURACY, value=accuracy)
    ws.cell(row=row_number, column=COLUMN_PRECISION, value=precision)
    ws.cell(row=row_number, column=COLUMN_RECALL, value=recall)
    ws.cell(row=row_number, column=COLUMN_F1, value=f1)
    # AUC для каждого класса и усредненный AUC
    y_true = np.array(test_labels)
    fpr, tpr, thresholds = roc_curve(y_true, probs)
    roc_auc = auc(fpr, tpr)
    ws.cell(row=row_number, column=COLUMN_AUC, value=roc_auc)
wb.save(f'content\\Сравнение оценок качества работы моделей RuBioBERT для задачи RE.xlsx')
print('Работа программы успешно завершина!')